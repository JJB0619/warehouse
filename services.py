from models import db, User, Warehouse, Product, StockIn, StockOut, Transfer
from flask_login import current_user
from openpyxl import Workbook
from datetime import datetime
import os

# ====================== 系统初始化：自动创建仓库+管理员 ======================
def init_system():
    # 自动创建4个固定仓库
    warehouse_names = ['后厨', '吧台', '二楼', '三楼']
    for name in warehouse_names:
        if not Warehouse.query.filter_by(name=name).first():
            wh = Warehouse(name=name)
            db.session.add(wh)
    
    # 自动创建管理员账号（默认：admin / 123456）
    if not User.query.filter_by(username='admin').first():
        admin = User(username='admin', is_admin=True)
        admin.set_password('123456')
        db.session.add(admin)
    
    db.session.commit()

# ====================== 权限校验 ======================
def is_admin_user():
    return current_user.is_authenticated and current_user.is_admin

# ====================== 库存预警 ======================
def get_stock_warning():
    from config import Config
    warning_list = []
    products = Product.query.all()
    for p in products:
        if p.stock < Config.STOCK_WARNING_THRESHOLD:
            wh = Warehouse.query.get(p.warehouse_id)
            warning_list.append({
                'id': p.id,
                'name': p.name,
                'warehouse': wh.name if wh else '未知',
                'stock': p.stock,
                'threshold': Config.STOCK_WARNING_THRESHOLD
            })
    return warning_list

# ====================== 出入库/调拨核心逻辑 ======================
def stock_in_op(product_id, warehouse_id, quantity, operator):
    product = Product.query.get(product_id)
    product.stock += quantity
    record = StockIn(
        product_id=product_id,
        warehouse_id=warehouse_id,
        quantity=quantity,
        operator=operator
    )
    db.session.add(record)
    db.session.commit()

def stock_out_op(product_id, warehouse_id, quantity, operator):
    product = Product.query.get(product_id)
    if product.stock < quantity:
        return False
    product.stock -= quantity
    record = StockOut(
        product_id=product_id,
        warehouse_id=warehouse_id,
        quantity=quantity,
        operator=operator
    )
    db.session.add(record)
    db.session.commit()
    return True

def transfer_op(product_id, from_wh_id, to_wh_id, quantity, operator):
    # 校验调出商品库存
    from_product = Product.query.filter_by(id=product_id, warehouse_id=from_wh_id).first()
    if not from_product or from_product.stock < quantity:
        return False

    # 扣减调出仓库库存
    from_product.stock -= quantity

    # 处理调入仓库：存在则累加，不存在则新建商品
    to_product = Product.query.filter_by(name=from_product.name, warehouse_id=to_wh_id).first()
    if to_product:
        to_product.stock += quantity
    else:
        new_product = Product(
            name=from_product.name,
            spec=from_product.spec,
            stock=quantity,
            warehouse_id=to_wh_id
        )
        db.session.add(new_product)

    # 记录调拨日志
    transfer_record = Transfer(
        product_id=product_id,
        from_warehouse_id=from_wh_id,
        to_warehouse_id=to_wh_id,
        quantity=quantity,
        operator=operator
    )
    db.session.add(transfer_record)
    db.session.commit()
    return True

# ====================== Excel全报表导出 ======================
def export_all_reports():
    wb = Workbook()
    
    # 1. 商品库存表
    ws_stock = wb.active
    ws_stock.title = "商品库存"
    ws_stock.append(['商品ID', '商品名称', '规格', '所属仓库', '当前库存', '创建时间'])
    for p in Product.query.all():
        wh = Warehouse.query.get(p.warehouse_id)
        ws_stock.append([
            p.id, p.name, p.spec,
            wh.name if wh else '未知',
            p.stock, p.create_time.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # 2. 入库记录表
    ws_in = wb.create_sheet("入库记录")
    ws_in.append(['记录ID', '商品名称', '仓库', '入库数量', '操作人', '操作时间'])
    for r in StockIn.query.all():
        p = Product.query.get(r.product_id)
        wh = Warehouse.query.get(r.warehouse_id)
        ws_in.append([
            r.id, p.name if p else '未知',
            wh.name if wh else '未知',
            r.quantity, r.operator,
            r.create_time.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # 3. 出库记录表
    ws_out = wb.create_sheet("出库记录")
    ws_out.append(['记录ID', '商品名称', '仓库', '出库数量', '操作人', '操作时间'])
    for r in StockOut.query.all():
        p = Product.query.get(r.product_id)
        wh = Warehouse.query.get(r.warehouse_id)
        ws_out.append([
            r.id, p.name if p else '未知',
            wh.name if wh else '未知',
            r.quantity, r.operator,
            r.create_time.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # 4. 调拨记录表
    ws_transfer = wb.create_sheet("调拨记录")
    ws_transfer.append(['记录ID', '商品名称', '调出仓库', '调入仓库', '调拨数量', '操作人', '操作时间'])
    for t in Transfer.query.all():
        p = Product.query.get(t.product_id)
        fwh = Warehouse.query.get(t.from_warehouse_id)
        twh = Warehouse.query.get(t.to_warehouse_id)
        ws_transfer.append([
            t.id, p.name if p else '未知',
            fwh.name if fwh else '未知',
            twh.name if twh else '未知',
            t.quantity, t.operator,
            t.create_time.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # 保存文件到静态目录
    report_dir = 'static/reports'
    os.makedirs(report_dir, exist_ok=True)
    filename = f"餐饮仓库全报表_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    file_path = os.path.join(report_dir, filename)
    wb.save(file_path)
    return file_path